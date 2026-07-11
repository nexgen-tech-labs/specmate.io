"""Parses uploaded Excel (.xlsx) and CSV files into row-level text chunks.

Location pointers use real spreadsheet row numbers (1-indexed, header row included
in the count), not re-numbered indices: xlsx chunks get "{sheet}/r.{N}", CSV chunks
get "r.{N}". The first row of each sheet/file is treated as a header row by default
(`has_header=True` — no per-upload override UI yet); each data row is rendered as
"Header: value | Header: value" pairs so column semantics survive into the text.

Messy real-world exports degrade gracefully rather than crash: merged cells read as
empty in all but their anchor cell, rows with more cells than headers keep the extra
values un-labelled, and empty rows are skipped entirely.
"""

from __future__ import annotations

import csv
import io
import zipfile

from openpyxl import load_workbook

from app.services.parsing.types import ParsedChunk


class SpreadsheetParseError(Exception):
    """Raised when an .xlsx/.csv file can't be parsed — corrupt file, wrong format,
    or no extractable (non-empty) rows."""


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_to_text(headers: list[str], cells: list[str]) -> str:
    """Renders one data row as "Header: value | Header: value" pairs, dropping empty
    cells. Cells beyond the header row's width are kept without a label."""
    parts: list[str] = []
    for i, cell in enumerate(cells):
        if not cell:
            continue
        header = headers[i] if i < len(headers) and headers[i] else None
        parts.append(f"{header}: {cell}" if header else cell)
    return " | ".join(parts)


def _rows_to_chunks(
    rows: list[tuple[int, list[str]]],
    section_prefix: str,
    order_start: int,
    has_header: bool,
) -> tuple[list[ParsedChunk], int]:
    """rows: (real_row_number, cell_strings) pairs, already stripped. Returns the
    chunks plus the next order value."""
    headers: list[str] = []
    chunks: list[ParsedChunk] = []
    order = order_start
    header_consumed = False

    for row_number, cells in rows:
        if not any(cells):
            continue
        if has_header and not header_consumed:
            headers = cells
            header_consumed = True
            continue

        text = _row_to_text(headers, cells)
        if not text:
            continue
        chunks.append(
            ParsedChunk(
                text=text,
                section_path=f"{section_prefix}r.{row_number}",
                order=order,
            )
        )
        order += 1

    return chunks, order


def parse_xlsx(file_bytes: bytes, has_header: bool = True) -> list[ParsedChunk]:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except (zipfile.BadZipFile, KeyError, ValueError, OSError) as exc:
        raise SpreadsheetParseError(
            f"Could not parse .xlsx file — it may be corrupt or not a valid Excel file: {exc}"
        ) from exc

    chunks: list[ParsedChunk] = []
    order = 0
    for sheet in workbook.worksheets:
        rows = [
            (row_number, [_cell_str(cell) for cell in row])
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1)
        ]
        sheet_chunks, order = _rows_to_chunks(
            rows, section_prefix=f"{sheet.title}/", order_start=order, has_header=has_header
        )
        chunks.extend(sheet_chunks)

    if not chunks:
        raise SpreadsheetParseError("Workbook contains no extractable rows.")
    return chunks


def _decode_csv_bytes(file_bytes: bytes) -> str:
    # utf-8-sig strips a BOM if present; fall back to latin-1 (never fails) for
    # legacy exports rather than rejecting the file outright.
    try:
        return file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        return file_bytes.decode("latin-1")


def parse_csv(file_bytes: bytes, has_header: bool = True) -> list[ParsedChunk]:
    text = _decode_csv_bytes(file_bytes)

    try:
        reader = csv.reader(io.StringIO(text))
        rows = [
            (row_number, [cell.strip() for cell in row])
            for row_number, row in enumerate(reader, start=1)
        ]
    except csv.Error as exc:
        raise SpreadsheetParseError(f"Could not parse CSV file: {exc}") from exc

    chunks, _ = _rows_to_chunks(rows, section_prefix="", order_start=0, has_header=has_header)

    if not chunks:
        raise SpreadsheetParseError("CSV contains no extractable rows.")
    return chunks
